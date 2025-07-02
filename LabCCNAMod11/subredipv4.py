import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def crear_tabla_subredes():
    """
    Crea una tabla de configuración de red similar a la imagen
    """
    
    # Datos para la tabla de subredes /26
    subredes_data = {
        'Dirección de subred': ['192.168.0.0', '192.168.0.64', '192.168.0.128', '192.168.0.192'],
        'Prefijo': ['/26', '/26', '/26', '/26'],
        'Máscara de subred': ['255.255.255.192', '255.255.255.192', '255.255.255.192', '255.255.255.192']
    }
    
    # Datos para la tabla de dispositivos
    dispositivos_data = {
        'Dispositivo': [
            'CustomerRouter', '', '', 
            'LAN_A Switch', 'LAN_B Switch', 
            'PC_A', 'PC_B', 
            'ISPRouter', '', 
            'ISPSwitch', 
            'Estación de trabajo ISP', 
            'Servidor del ISP'
        ],
        'Interfaz': [
            'G0/0', 'G0/1', 'S0/1/0',
            'VLAN1', 'VLAN1',
            'NIC', 'NIC',
            'G0/0', 'S0/1/0',
            'VLAN1',
            'NIC',
            'NIC'
        ],
        'Dirección IP': [
            '', '', '209.165.201.2',
            '', '',
            '', '',
            '209.165.200.225', '209.165.201.1',
            '209.165.200.226',
            '209.165.200.235',
            '209.165.200.240'
        ],
        'Máscara de subred': [
            '', '', '255.255.255.252',
            '', '',
            '', '',
            '255.255.255.224', '255.255.255.252',
            '255.255.255.224',
            '255.255.255.224',
            '255.255.255.224'
        ],
        'Gateway predeterminado': [
            'No corresponde', '', '',
            '', '',
            '', '',
            'N/D', '',
            '209.165.200.225',
            '209.165.200.225',
            '209.165.200.225'
        ]
    }
    
    return subredes_data, dispositivos_data

def formatear_excel(archivo='configuracion_red.xlsx'):
    """
    Crea un archivo Excel con formato similar a la imagen
    """
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Configuración de Red"
    
    # Obtener datos
    subredes_data, dispositivos_data = crear_tabla_subredes()
    
    # Configurar estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título principal para información de máscara
    ws['A1'] = '/26'
    ws['B1'] = '11111111.11111111.11111111.11000000'
    ws['C1'] = '255.255.255.192'
    
    # Aplicar formato al título
    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws[f'{col}1'].border = border
    
    # Tabla de subredes (empezar en fila 3)
    row_start = 3
    
    # Headers de subredes
    headers_subredes = ['Dirección de subred', 'Prefijo', 'Máscara de subred']
    for col, header in enumerate(headers_subredes, 1):
        cell = ws.cell(row=row_start, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos de subredes
    for row_idx, row_data in enumerate(zip(
        subredes_data['Dirección de subred'],
        subredes_data['Prefijo'],
        subredes_data['Máscara de subred']
    ), row_start + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # Tabla de dispositivos (empezar después de subredes + espacio)
    device_start = row_start + len(subredes_data['Dirección de subred']) + 3
    
    # Headers de dispositivos
    headers_dispositivos = ['Dispositivo', 'Interfaz', 'Dirección IP', 'Máscara de subred', 'Gateway predeterminado']
    for col, header in enumerate(headers_dispositivos, 1):
        cell = ws.cell(row=device_start, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos de dispositivos
    for row_idx, row_data in enumerate(zip(
        dispositivos_data['Dispositivo'],
        dispositivos_data['Interfaz'],
        dispositivos_data['Dirección IP'],
        dispositivos_data['Máscara de subred'],
        dispositivos_data['Gateway predeterminado']
    ), device_start + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 1:  # Columna de dispositivo
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
    
    # Ajustar ancho de columnas
    column_widths = [20, 15, 18, 20, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Guardar archivo
    wb.save(archivo)
    print(f"Archivo guardado como: {archivo}")

def generar_tabla_subredes_completa(red_base="192.168.0.0", prefijo_original=24, nuevo_prefijo=26):
    """
    Genera tabla completa de subredes basada en parámetros
    """
    
    # Calcular número de subredes
    bits_prestados = nuevo_prefijo - prefijo_original
    num_subredes = 2 ** bits_prestados
    hosts_por_subred = 2 ** (32 - nuevo_prefijo) - 2
    
    # Calcular el salto entre subredes
    salto = 2 ** (32 - nuevo_prefijo)
    
    # Generar subredes
    subredes = []
    base_parts = red_base.split('.')
    base_int = int(base_parts[3])
    
    for i in range(num_subredes):
        subnet_num = base_int + (i * salto)
        subnet_addr = f"{base_parts[0]}.{base_parts[1]}.{base_parts[2]}.{subnet_num}"
        
        # Calcular primera y última IP utilizables
        primera_ip = subnet_num + 1
        ultima_ip = subnet_num + salto - 2
        broadcast = subnet_num + salto - 1
        
        subredes.append({
            'Subred': i + 1,
            'Dirección de red': subnet_addr,
            'Prefijo': f'/{nuevo_prefijo}',
            'Primera IP utilizable': f"{base_parts[0]}.{base_parts[1]}.{base_parts[2]}.{primera_ip}",
            'Última IP utilizable': f"{base_parts[0]}.{base_parts[1]}.{base_parts[2]}.{ultima_ip}",
            'Dirección broadcast': f"{base_parts[0]}.{base_parts[1]}.{base_parts[2]}.{broadcast}",
            'Hosts utilizables': hosts_por_subred
        })
    
    # Crear DataFrame y exportar
    df = pd.DataFrame(subredes)
    df.to_excel('subredes_completas.xlsx', index=False)
    print(f"Tabla completa de subredes guardada como: subredes_completas.xlsx")
    print(f"Total de subredes generadas: {num_subredes}")
    print(f"Hosts por subred: {hosts_por_subred}")

if __name__ == "__main__":
    print("Generando tabla de configuración de red...")
    formatear_excel()
    
    print("\nGenerando tabla completa de subredes...")
    generar_tabla_subredes_completa()
    
    print("\n¡Archivos Excel generados exitosamente!")
    print("- configuracion_red.xlsx: Tabla de configuración de dispositivos")
    print("- subredes_completas.xlsx: Tabla completa de subredes /26")